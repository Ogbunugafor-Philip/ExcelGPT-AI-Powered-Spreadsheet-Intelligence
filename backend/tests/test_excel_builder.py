"""Tests for the Phase 5 Excel generation engine and the /download endpoint."""

import openpyxl
import pytest
from fastapi.testclient import TestClient

import config
import main
from schemas.computation_schema import (
    AnalysisSheet,
    Chart,
    ComputationOutput,
    ConditionalFormatRule,
    DataSheet,
    ExecutiveSummary,
    ForecastSheet,
    KpiCard,
    Metric,
)
from services.excel_builder import ExcelBuilder


def sample_output(session_id="builder-test") -> ComputationOutput:
    return ComputationOutput(
        session_id=session_id,
        version=1,
        executive_summary=ExecutiveSummary(
            title="Top 5 Branches by Deposit Volume",
            period="Jan 2026 – Jun 2026",
            data_source="branch_performance.xlsx (24 rows)",
            kpi_cards=[
                KpiCard(label="Top by Deposits", value="Lagos Island", change="₦187.5M", direction="up"),
                KpiCard(label="Avg Growth", value="4.57%", change="20↑ / 0↓", direction="up"),
            ],
        ),
        data_sheet=DataSheet(
            columns=["Rank", "branch", "deposits_ngn", "growth_pct", "month"],
            rows=[
                [1, "Lagos Island", 187500000, 25.0, "2026-06"],
                [2, "Abuja", 150000000, -5.0, "2026-05"],
                [3, "Ikeja", 122000000, 8.0, "2026-04"],
            ],
            conditional_formatting=[ConditionalFormatRule(column="growth_pct", rule="value < 0", color="EF4444")],
        ),
        analysis_sheet=AnalysisSheet(
            metrics=[Metric(label="Total Deposits", value="₦2.55B", formula_used="sum over 4 groups")],
            rankings=[
                {"Rank": 1, "branch": "Lagos Island", "deposits_ngn": 187500000},
                {"Rank": 2, "branch": "Abuja", "deposits_ngn": 150000000},
                {"Rank": 3, "branch": "Ikeja", "deposits_ngn": 122000000},
            ],
            growth_table=[
                {"branch": "Abuja", "period": "2026-02", "growth_pct": 5.0, "direction": "up"},
                {"branch": "Abuja", "period": "2026-03", "growth_pct": -2.0, "direction": "down"},
            ],
        ),
        charts=[Chart(chart_id="op_4", chart_type="bar", title="Deposits by Branch", image_path="", recharts_data=[])],
        forecast_sheet=ForecastSheet(
            historical=[{"period": "2026-01", "value": 100}, {"period": "2026-02", "value": 110}],
            projected=[{"period": "2026-03", "value": 120}],
            confidence_upper=[{"period": "2026-03", "value": 130}],
            confidence_lower=[{"period": "2026-03", "value": 110}],
            assumptions=["ExponentialSmoothing (trend=add)", "3 periods ahead", "95% confidence"],
        ),
    )


def test_build_returns_valid_xlsx_path():
    path = ExcelBuilder().build(sample_output(), "builder-test")
    assert path.endswith(".xlsx")
    wb = openpyxl.load_workbook(path)  # raises if not a valid workbook
    assert wb is not None


def test_workbook_contains_expected_sheets():
    path = ExcelBuilder().build(sample_output(), "builder-test")
    wb = openpyxl.load_workbook(path)
    assert wb.sheetnames == ["Executive Summary", "Data", "Analysis", "Charts", "Forecast"]


def test_executive_summary_has_kpi_cards():
    path = ExcelBuilder().build(sample_output(), "builder-test")
    ws = openpyxl.load_workbook(path)["Executive Summary"]
    text = {cell.value for row in ws.iter_rows() for cell in row if isinstance(cell.value, str)}
    assert "ExcelGPT" in text                       # logo
    assert "Top by Deposits" in text                # KPI label
    assert "Lagos Island" in text                   # KPI value


def test_data_sheet_has_frozen_panes_and_filter():
    path = ExcelBuilder().build(sample_output(), "builder-test")
    ws = openpyxl.load_workbook(path)["Data"]
    assert ws.freeze_panes == "A2"
    assert ws.auto_filter.ref is not None


def test_empty_sheets_are_skipped():
    output = sample_output()
    output.charts = []
    output.forecast_sheet = ForecastSheet()  # empty -> skipped
    path = ExcelBuilder().build(output, "builder-empty")
    wb = openpyxl.load_workbook(path)
    assert "Charts" not in wb.sheetnames
    assert "Forecast" not in wb.sheetnames
    assert "Executive Summary" in wb.sheetnames


def test_download_endpoint_returns_xlsx_stream():
    client = TestClient(main.app)
    session_id = "download-test"
    token = "tok-download-test"
    main.session_manager.create_session(
        file_path="unused.xlsx",
        intelligence_brief={"filename": "branch_performance.xlsx"},
        sheet_data={},
        session_id=session_id,
    )
    main.session_manager.get_session(session_id)["downloads"] = {token: sample_output(session_id).model_dump()}

    response = client.get(f"/download/{token}")
    assert response.status_code == 200
    assert response.headers["content-type"] == config.XLSX_MEDIA_TYPE
    assert "attachment" in response.headers["content-disposition"]
    assert response.content[:2] == b"PK"  # .xlsx is a zip archive


def test_download_unknown_token_returns_404():
    client = TestClient(main.app)
    assert client.get("/download/does-not-exist").status_code == 404
