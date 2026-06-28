from __future__ import annotations

import os
import tempfile
from pathlib import Path
from typing import Any

import pandas as pd
from fastapi import HTTPException, UploadFile
from openpyxl import load_workbook

import config


class FileValidator:
    """Validate uploaded workbook files before ingestion."""

    def validate(self, file: UploadFile) -> None:
        if not file.filename:
            raise HTTPException(status_code=400, detail="No file was provided.")

        suffix = Path(file.filename).suffix.lower()
        if suffix not in config.ALLOWED_EXTENSIONS:
            raise HTTPException(
                status_code=400,
                detail="Unsupported file type. Please upload an Excel workbook (.xlsx or .xls).",
            )

        temp_handle = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
        temp_path = Path(temp_handle.name)
        temp_handle.close()

        size = 0
        try:
            with temp_path.open("wb") as buffer:
                file.file.seek(0)
                while chunk := file.file.read(1024 * 1024):
                    size += len(chunk)
                    buffer.write(chunk)
            file.file.seek(0)
        except Exception as exc:  # pragma: no cover - defensive
            raise HTTPException(status_code=400, detail="The uploaded file could not be read.") from exc

        if size > config.MAX_FILE_SIZE:
            temp_path.unlink(missing_ok=True)
            raise HTTPException(status_code=413, detail=f"File exceeds the {config.MAX_FILE_SIZE_MB}MB limit.")

        try:
            if suffix == ".xlsx":
                load_workbook(temp_path, read_only=True, data_only=True)
            else:
                try:
                    load_workbook(temp_path, read_only=True, data_only=True)
                except Exception:
                    pd.ExcelFile(temp_path, engine="xlrd")
        except Exception as exc:
            temp_path.unlink(missing_ok=True)
            raise HTTPException(
                status_code=400,
                detail="The uploaded workbook could not be read. Please confirm it is a valid Excel file.",
            ) from exc

        if not self._has_non_empty_sheet(temp_path):
            temp_path.unlink(missing_ok=True)
            raise HTTPException(status_code=400, detail="The workbook does not contain any usable sheets.")

        setattr(file, "_validated_temp_path", temp_path)

    def _has_non_empty_sheet(self, temp_path: Path) -> bool:
        try:
            excel_file = pd.ExcelFile(temp_path)
        except Exception:
            return False
        for sheet_name in excel_file.sheet_names:
            try:
                first_sheet = pd.read_excel(temp_path, sheet_name=sheet_name, header=None)
            except Exception:
                continue
            if first_sheet.empty:
                continue
            if not first_sheet.dropna(how="all").empty:
                return True
        return False
