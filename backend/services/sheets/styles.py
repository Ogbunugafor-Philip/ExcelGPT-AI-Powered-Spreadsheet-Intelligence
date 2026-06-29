"""
Shared styling primitives for the Excel sheet builders.

Centralises the ExcelGPT colour system and the small openpyxl helpers (fills,
fonts, borders, alignment, column/canvas helpers, column-type detection) so the
five sheet builders stay focused on layout.

The workbook uses a clean, light "senior analyst" theme: a navy header band,
white / very-light-blue alternating rows, dark readable text, subtle green/red
tints for positive/negative values and gold for the top rank.
"""

from __future__ import annotations

from typing import Any

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# -- ExcelGPT colour system (aRGB hex, no leading '#') ----------------------
# World-class light theme palette (FIX 6).
HEADER_BG = "1F3864"       # dark navy — header row background
HEADER_FONT = "FFFFFF"     # white — header text
ROW_ALT = "F0F4FA"         # very light blue-grey — alternating data rows
ROW_MAIN = "FFFFFF"        # white — main data rows
POSITIVE_BG = "E8F5E9"     # light green — positive values
NEGATIVE_BG = "FFEBEE"     # light red — negative values
GOLD_BG = "FFF8E1"         # light gold — rank 1
SECTION_HEADER = "2E75B6"  # medium blue — section title band
SECTION_FONT = "FFFFFF"    # white — section title text
BORDER_COLOR = "D5DBE5"    # soft grey — cell borders
TEXT_DARK = "1F2937"       # near-black slate — primary data text
TEXT_MUTED = "6B7280"      # grey — secondary / muted text
POSITIVE_TEXT = "1B7F3B"   # green text
NEGATIVE_TEXT = "C0392B"   # red text

# Legacy / accent colours (kept for backwards compatibility & charts).
NAVY = HEADER_BG
NAVY_LIGHT = "111827"
BLUE_ELECTRIC = "2563EB"
EMERALD = "10B981"
AMBER = "F59E0B"
RED_ALERT = "EF4444"
GOLD = "D97706"
WHITE = "FFFFFF"
GREY_LIGHT = "F3F4F6"
GREY_TEXT = TEXT_MUTED
TEXT_PRIMARY = TEXT_DARK
BRONZE = "B45309"          # rank-3 medal
GREEN_TINT = POSITIVE_BG   # positive growth cell tint
RED_TINT = NEGATIVE_BG     # negative growth cell tint

DEFAULT_FONT = "Calibri"

DATE_KEYWORDS = ("date", "month", "period", "year", "day", "quarter", "time")
PCT_KEYWORDS = ("pct", "percent", "growth", "rate", "margin", "ratio", "%", "variance", "change")
CURRENCY_KEYWORDS = (
    "ngn", "naira", "deposit", "loan", "amount", "balance", "revenue", "salary",
    "income", "cost", "price", "value", "fee", "asset", "turnover", "sales", "profit", "spend", "budget",
    "target", "₦",
)

# -- number formats ---------------------------------------------------------
# Naira amounts. Whole-number form is the default senior-analyst look
# (₦573,750,000); the decimal form is available where kobo precision matters.
NAIRA_FORMAT = '"₦"#,##0'
NAIRA_DECIMAL_FORMAT = '"₦"#,##0.00'
NUMBER_FORMAT = '#,##0'             # large integers with comma separators
DECIMAL_FORMAT = '#,##0.00'         # decimals with comma separators
# NOTE: percentages in this engine are stored as WHOLE numbers (20.0 == 20%),
# so the format appends a literal '%' rather than using Excel's '0.0%' (which
# would multiply by 100 and render 20.0 as 2000%).
PERCENT_FORMAT = '#,##0.0"%"'       # 24.41 -> 24.4%
PERCENT_INT_FORMAT = '#,##0"%"'     # 24.41 -> 24%
DATE_FORMAT = "DD-MMM-YYYY"

# Back-compat alias used by the global named style in excel_builder.
CURRENCY_FORMAT = NAIRA_DECIMAL_FORMAT


def fill(color: str) -> PatternFill:
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def font(size: int = 11, bold: bool = False, italic: bool = False, color: str = TEXT_DARK) -> Font:
    return Font(name=DEFAULT_FONT, size=size, bold=bold, italic=italic, color=color)


def side(color: str = BORDER_COLOR, style: str = "thin") -> Side:
    return Side(style=style, color=color)


def box(color: str = BORDER_COLOR, style: str = "thin") -> Border:
    edge = side(color, style)
    return Border(left=edge, right=edge, top=edge, bottom=edge)


def header_border() -> Border:
    """Thin sides + a medium navy bottom rule under the header row."""
    thin = side(BORDER_COLOR, "thin")
    return Border(left=thin, right=thin, top=thin, bottom=side(HEADER_BG, "medium"))


CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")


def style_cell(cell, *, value=None, font_=None, fill_=None, align=None, border=None, number_format=None):
    """Apply a bundle of styles to a single cell in one call."""
    if value is not None:
        cell.value = value
    if font_ is not None:
        cell.font = font_
    if fill_ is not None:
        cell.fill = fill_
    if align is not None:
        cell.alignment = align
    if border is not None:
        cell.border = border
    if number_format is not None:
        cell.number_format = number_format
    return cell


def set_column_widths(ws, width: float, first: int = 1, last: int = 8) -> None:
    for index in range(first, last + 1):
        ws.column_dimensions[get_column_letter(index)].width = width


def autofit_columns(ws, columns: list[Any], rows: list[list[Any]], *, first: int = 1,
                    minimum: float = 10.0, maximum: float = 45.0, padding: float = 1.3) -> None:
    """Auto-fit each column to the widest of its header / first-100 values.

    width = max(len(header), max(len(str(value)) for the first 100 rows)) * padding,
    clamped to [minimum, maximum]. Apply AFTER all data is written.
    """
    for offset, col in enumerate(columns):
        header = column_name(col)
        max_len = len(str(header))
        for row in rows[:100]:
            if offset < len(row):
                value = row[offset]
                if value is not None:
                    max_len = max(max_len, len(_display_len(value)))
        width = max(minimum, min(maximum, max_len * padding))
        ws.column_dimensions[get_column_letter(first + offset)].width = width


def _display_len(value: Any) -> str:
    """Approximate the rendered text length of a value (commas widen numbers)."""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, int):
        return f"{value:,}"
    if isinstance(value, float):
        return f"{value:,.2f}"
    return str(value)


def paint_canvas(ws, max_row: int, max_col: int, color: str = ROW_MAIN) -> None:
    """Fill every still-unfilled cell in the box so the sheet reads as one canvas."""
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            if cell.fill is None or cell.fill.patternType is None:
                cell.fill = fill(color)


def hide_gridlines(ws) -> None:
    ws.sheet_view.showGridLines = False


# -- column-name / value type detection -------------------------------------

def column_name(column: Any) -> str:
    """Columns may be plain strings or {'name': ..., 'type': ...} dicts."""
    if isinstance(column, dict):
        return str(column.get("name", ""))
    return str(column)


def detect_type(name: str, values: list[Any]) -> str:
    """Classify a column as currency | percentage | date | number | text.

    Explicit symbols win first so that 'Variance (₦)' reads as currency while
    'Variance (%)' reads as percentage, even though both share the 'variance'
    keyword.
    """
    lowered = str(name).lower()
    if "%" in lowered or "percent" in lowered:
        return "percentage"
    if "₦" in lowered or "ngn" in lowered or "naira" in lowered:
        return "currency"
    # word-based percentage hints (growth, rate, variance, margin, …) win over
    # the broad currency keyword list.
    if any(k in lowered for k in PCT_KEYWORDS):
        return "percentage"
    if any(k in lowered for k in CURRENCY_KEYWORDS):
        return "currency"
    if any(k in lowered for k in DATE_KEYWORDS):
        return "date"

    non_null = [v for v in values if v is not None and str(v).strip() != ""]
    if not non_null:
        return "text"
    numeric = sum(1 for v in non_null if _is_number(v))
    if numeric / len(non_null) >= 0.8:
        return "number"
    if _date_ratio(non_null) >= 0.8:
        return "date"
    return "text"


def number_format_for(name: str, value: Any) -> str | None:
    """Pick the openpyxl number format for a cell from its column name + value.

    - currency names (₦, NGN, Deposit, Loan, Revenue, Target, Amount, …) -> Naira
    - percentage names (%, Rate, Growth, Variance, …)                    -> percent
    - integer values > 999 with no currency hint                        -> #,##0
    - float values                                                      -> #,##0.00
    """
    col_type = detect_type(name, [value])
    if col_type == "currency":
        return NAIRA_DECIMAL_FORMAT if isinstance(value, float) and not float(value).is_integer() else NAIRA_FORMAT
    if col_type == "percentage":
        return PERCENT_FORMAT
    if isinstance(value, bool):
        return None
    if isinstance(value, int) and abs(value) > 999:
        return NUMBER_FORMAT
    if isinstance(value, float):
        return DECIMAL_FORMAT if not value.is_integer() else NUMBER_FORMAT
    return None


def _is_number(value: Any) -> bool:
    if isinstance(value, bool):
        return False
    if isinstance(value, (int, float)):
        return True
    if isinstance(value, str):
        try:
            float(value.replace(",", "").replace("₦", "").replace("%", "").strip())
            return True
        except ValueError:
            return False
    return False


def _date_ratio(values: list[Any]) -> float:
    try:
        parsed = pd.to_datetime([str(v) for v in values], errors="coerce")
    except Exception:  # noqa: BLE001
        return 0.0
    return float(parsed.notna().mean())
