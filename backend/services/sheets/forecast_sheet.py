"""Forecast sheet — historical vs projected with confidence bounds and assumptions."""

from __future__ import annotations

import os
from typing import Any

from openpyxl.drawing.image import Image as XLImage

from . import styles as S

HEADERS = ["Period", "Historical", "Forecast", "Upper Bound", "Lower Bound"]


class ForecastSheet:
    @staticmethod
    def has_content(forecast: dict[str, Any] | None) -> bool:
        if not forecast:
            return False
        return bool(forecast.get("historical") or forecast.get("projected"))

    def build(self, ws, forecast: dict[str, Any]) -> None:
        S.hide_gridlines(ws)
        ws.sheet_properties.tabColor = S.GREY_TEXT
        S.set_column_widths(ws, 18, first=1, last=5)

        # Title
        ws.merge_cells("A1:E1")
        S.style_cell(ws["A1"], value="FORECAST ANALYSIS", font_=S.font(14, bold=True, color=S.BLUE_ELECTRIC), align=S.LEFT)

        # Table header
        header_row = 3
        border = S.box(S.BLUE_ELECTRIC, "thin")
        for idx, head in enumerate(HEADERS, start=1):
            S.style_cell(ws.cell(row=header_row, column=idx), value=head,
                         font_=S.font(11, bold=True, color=S.WHITE), fill_=S.fill(S.BLUE_ELECTRIC),
                         align=S.CENTER, border=border)
        ws.freeze_panes = "A2"

        row = header_row + 1
        row = self._historical_rows(ws, forecast.get("historical", []), row)
        row = self._forecast_rows(ws, forecast, row)

        row = self._assumptions(ws, forecast.get("assumptions", []), row + 1)
        self._maybe_chart(ws, forecast, row + 1)

        S.paint_canvas(ws, max_row=max(row + 1, 4), max_col=5, color=S.NAVY)

    def _historical_rows(self, ws, historical, row):
        for point in historical or []:
            S.style_cell(ws.cell(row=row, column=1), value=point.get("period"),
                         font_=S.font(10, color=S.WHITE), fill_=S.fill(S.NAVY), align=S.LEFT)
            S.style_cell(ws.cell(row=row, column=2), value=self._num(point.get("value")),
                         font_=S.font(10, color=S.WHITE), fill_=S.fill(S.NAVY), align=S.RIGHT, number_format='#,##0.00')
            for col in (3, 4, 5):
                S.style_cell(ws.cell(row=row, column=col), value="-",
                             font_=S.font(9, color=S.GREY_TEXT), fill_=S.fill(S.NAVY), align=S.CENTER)
            row += 1
        return row

    def _forecast_rows(self, ws, forecast, row):
        projected = forecast.get("projected", []) or []
        upper = {p.get("period"): p.get("value") for p in forecast.get("confidence_upper", []) or []}
        lower = {p.get("period"): p.get("value") for p in forecast.get("confidence_lower", []) or []}
        for point in projected:
            period = point.get("period")
            S.style_cell(ws.cell(row=row, column=1), value=period,
                         font_=S.font(10, italic=True, color=S.BLUE_ELECTRIC), fill_=S.fill(S.NAVY_LIGHT), align=S.LEFT)
            S.style_cell(ws.cell(row=row, column=2), value="-",
                         font_=S.font(9, color=S.GREY_TEXT), fill_=S.fill(S.NAVY_LIGHT), align=S.CENTER)
            S.style_cell(ws.cell(row=row, column=3), value=self._num(point.get("value")),
                         font_=S.font(10, italic=True, color=S.BLUE_ELECTRIC), fill_=S.fill(S.NAVY_LIGHT),
                         align=S.RIGHT, number_format='#,##0.00')
            S.style_cell(ws.cell(row=row, column=4), value=self._num(upper.get(period)),
                         font_=S.font(9, color=S.GREY_TEXT), fill_=S.fill(S.NAVY_LIGHT), align=S.RIGHT, number_format='#,##0.00')
            S.style_cell(ws.cell(row=row, column=5), value=self._num(lower.get(period)),
                         font_=S.font(9, color=S.GREY_TEXT), fill_=S.fill(S.NAVY_LIGHT), align=S.RIGHT, number_format='#,##0.00')
            row += 1
        return row

    def _assumptions(self, ws, assumptions, row):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        S.style_cell(ws.cell(row=row, column=1), value="MODEL ASSUMPTIONS",
                     font_=S.font(11, bold=True, color=S.GREY_TEXT), align=S.LEFT)
        row += 1
        for assumption in assumptions or ["No assumptions recorded."]:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
            S.style_cell(ws.cell(row=row, column=1), value=f"•  {assumption}",
                         font_=S.font(10, color=S.WHITE), align=S.LEFT)
            row += 1
        return row

    def _maybe_chart(self, ws, forecast, row):
        path = forecast.get("chart_image_path")
        if path and os.path.exists(path):
            try:
                img = XLImage(path)
                img.width, img.height = 480, 300
                ws.add_image(img, f"A{row + 1}")
            except Exception:  # noqa: BLE001
                pass

    @staticmethod
    def _num(value):
        if value is None:
            return "-"
        try:
            return float(value)
        except (TypeError, ValueError):
            return value
