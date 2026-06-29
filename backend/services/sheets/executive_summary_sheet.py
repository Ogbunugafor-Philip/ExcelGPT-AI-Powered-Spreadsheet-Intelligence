"""Executive Summary sheet — branded header + question-driven KPI cards (light theme)."""

from __future__ import annotations

from typing import Any

from openpyxl.utils import get_column_letter

from . import styles as S

CARDS_PER_ROW = 4
ARROWS = {"up": ("▲", S.POSITIVE_TEXT), "down": ("▼", S.NEGATIVE_TEXT), "neutral": ("▬", S.TEXT_MUTED)}


class ExecutiveSummarySheet:
    def build(self, ws, summary: dict[str, Any]) -> None:
        S.hide_gridlines(ws)
        ws.sheet_properties.tabColor = S.SECTION_HEADER
        S.set_column_widths(ws, 15, first=1, last=12)

        # Row 1 — brand logo on the navy header band
        ws.merge_cells("A1:L1")
        S.style_cell(ws["A1"], value="ExcelGPT", font_=S.font(20, bold=True, color=S.HEADER_FONT),
                     fill_=S.fill(S.HEADER_BG), align=S.LEFT)
        for col in range(2, 13):
            ws.cell(row=1, column=col).fill = S.fill(S.HEADER_BG)

        # Row 2 — report title (reflects the actual question asked)
        ws.merge_cells("A2:L2")
        S.style_cell(ws["A2"], value=summary.get("title", "ExcelGPT Report"),
                     font_=S.font(16, bold=True, color=S.HEADER_BG), align=S.LEFT)

        # Row 3 — data source | period
        ws.merge_cells("A3:L3")
        meta = f"Data Source: {summary.get('data_source', '—')}    |    Period: {summary.get('period', '—')}"
        S.style_cell(ws["A3"], value=meta, font_=S.font(10, color=S.TEXT_MUTED), align=S.LEFT)

        ws.row_dimensions[1].height = 32
        ws.row_dimensions[2].height = 24

        # Row 5+ — KPI cards (4 per row, 3-row pitch + spacer)
        cards = summary.get("kpi_cards", []) or []
        last_row = 4
        for index, card in enumerate(cards):
            col_block = index % CARDS_PER_ROW
            row_block = index // CARDS_PER_ROW
            start_col = 1 + col_block * 3          # A, D, G, J
            start_row = 5 + row_block * 4
            last_row = max(last_row, self._render_card(ws, card, start_row, start_col))

        if not cards:
            S.style_cell(ws["A5"], value="No KPIs computed for this report.",
                         font_=S.font(11, italic=True, color=S.TEXT_MUTED))
            last_row = 5

        S.paint_canvas(ws, max_row=last_row + 1, max_col=12, color=S.ROW_MAIN)

    def _render_card(self, ws, card: dict[str, Any], start_row: int, start_col: int) -> int:
        left = get_column_letter(start_col)
        right = get_column_letter(start_col + 1)
        border = S.box(S.SECTION_HEADER, "thin")
        card_bg = S.ROW_ALT

        # Label
        ws.merge_cells(f"{left}{start_row}:{right}{start_row}")
        S.style_cell(ws.cell(row=start_row, column=start_col),
                     value=card.get("label", ""), font_=S.font(9, bold=True, color=S.TEXT_MUTED),
                     fill_=S.fill(card_bg), align=S.LEFT)

        # Value
        ws.merge_cells(f"{left}{start_row + 1}:{right}{start_row + 1}")
        S.style_cell(ws.cell(row=start_row + 1, column=start_col),
                     value=card.get("value", ""), font_=S.font(17, bold=True, color=S.HEADER_BG),
                     fill_=S.fill(card_bg), align=S.LEFT)
        ws.row_dimensions[start_row + 1].height = 26

        # Change with direction arrow
        direction = card.get("direction", "neutral")
        arrow, color = ARROWS.get(direction, ARROWS["neutral"])
        change_text = f"{arrow} {card.get('change', '')}".strip()
        ws.merge_cells(f"{left}{start_row + 2}:{right}{start_row + 2}")
        S.style_cell(ws.cell(row=start_row + 2, column=start_col),
                     value=change_text, font_=S.font(10, color=color),
                     fill_=S.fill(card_bg), align=S.LEFT)

        # Border around the 2x3 card block
        for r in range(start_row, start_row + 3):
            for c in (start_col, start_col + 1):
                ws.cell(row=r, column=c).border = border

        return start_row + 2
