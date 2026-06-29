"""Analysis sheet — question-driven KEY METRICS, the PRIMARY RESULTS table
(rankings with medals or a colour-coded growth/variance table) and a written
INSIGHTS section. Everything is built from the actual computation output."""

from __future__ import annotations

from typing import Any

from openpyxl.utils import get_column_letter

from . import styles as S

# Rank medals (light theme): gold / silver / bronze tints.
MEDALS = {1: S.GOLD_BG, 2: "ECEFF4", 3: "F3E5D8"}
SECTION_SPAN = 8

# Raw computation keys -> human-readable display headers (PROBLEM 1).
_HUMAN_HEADERS = {
    "variance_pct": "Variance %",
    "variance_amount": "Variance (₦)",
    "growth_pct": "Growth Rate",
    "growth_rate": "Growth Rate",
    "status": "Status",
    "direction": "Trend",
    "period": "Period",
    "deposits_ngn": "Deposits (₦)",
    "rank": "Rank",
}


def humanize_header(col: Any) -> str:
    """variance_pct -> Variance %, status -> Status; leave already-display names
    (those carrying symbols / spaces) untouched; otherwise Title-Case + spaces."""
    key = str(col)
    mapped = _HUMAN_HEADERS.get(key.lower())
    if mapped:
        return mapped
    if any(ch in key for ch in ("→", "(", "%", "₦", " ")):
        return key
    return key.replace("_", " ").title()


class AnalysisSheet:
    def build(self, ws, analysis: dict[str, Any]) -> None:
        S.hide_gridlines(ws)
        ws.sheet_properties.tabColor = S.GOLD

        row = 1
        row = self._metrics(ws, analysis.get("metrics", []), row)
        row = self._rankings(ws, analysis.get("rankings", []), row + 1)
        row = self._growth(ws, analysis.get("growth_table", []), row + 1)
        row = self._insights(ws, analysis.get("insights", []), row + 1)

        # Widen the label column; keep the rest comfortable.
        ws.column_dimensions["A"].width = 28
        S.set_column_widths(ws, 18, first=2, last=SECTION_SPAN)
        ws.freeze_panes = "A2"
        S.paint_canvas(ws, max_row=max(row, 2), max_col=SECTION_SPAN, color=S.ROW_MAIN)

    def _section_header(self, ws, text, row, span=SECTION_SPAN):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
        S.style_cell(ws.cell(row=row, column=1), value=text,
                     font_=S.font(12, bold=True, color=S.SECTION_FONT), fill_=S.fill(S.SECTION_HEADER), align=S.LEFT)
        for col in range(2, span + 1):
            ws.cell(row=row, column=col).fill = S.fill(S.SECTION_HEADER)
        ws.row_dimensions[row].height = 20
        return row + 1

    # -- metrics ------------------------------------------------------------

    def _metrics(self, ws, metrics, row):
        row = self._section_header(ws, "KEY METRICS", row)
        if not metrics:
            S.style_cell(ws.cell(row=row, column=1), value="No metrics computed.",
                         font_=S.font(10, italic=True, color=S.TEXT_MUTED), fill_=S.fill(S.ROW_MAIN))
            return row + 1
        border = S.box(S.BORDER_COLOR, "thin")
        for i, metric in enumerate(metrics):
            bg = S.ROW_ALT if i % 2 == 0 else S.ROW_MAIN
            S.style_cell(ws.cell(row=row, column=1), value=metric.get("label", ""),
                         font_=S.font(11, bold=True, color=S.TEXT_DARK), fill_=S.fill(bg), align=S.LEFT, border=border)
            S.style_cell(ws.cell(row=row, column=2), value=metric.get("value", ""),
                         font_=S.font(11, bold=True, color=S.SECTION_HEADER), fill_=S.fill(bg), align=S.RIGHT, border=border)
            ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=SECTION_SPAN)
            S.style_cell(ws.cell(row=row, column=3), value=metric.get("formula_used", ""),
                         font_=S.font(9, italic=True, color=S.TEXT_MUTED), fill_=S.fill(bg), align=S.LEFT)
            ws.row_dimensions[row].height = 16
            row += 1
        return row

    # -- rankings -----------------------------------------------------------

    def _rankings(self, ws, rankings, row):
        if not rankings:
            return row
        row = self._section_header(ws, "PERFORMANCE RANKINGS", row)
        columns = [c for c in rankings[0].keys() if c != "direction"]
        row = self._table_header(ws, columns, row)
        border = S.box(S.BORDER_COLOR, "thin")
        for record in rankings:
            rank = record.get("Rank")
            medal = MEDALS.get(rank)
            bg = medal or (S.ROW_ALT if row % 2 == 0 else S.ROW_MAIN)
            bold = bool(medal)
            for c_index, col in enumerate(columns, start=1):
                value = record.get(col)
                S.style_cell(ws.cell(row=row, column=c_index), value=self._fmt(value),
                             font_=S.font(10, bold=bold, color=S.TEXT_DARK), fill_=S.fill(bg),
                             align=S.RIGHT if isinstance(value, (int, float)) and not isinstance(value, bool) else S.LEFT,
                             border=border, number_format=S.number_format_for(col, value))
            ws.row_dimensions[row].height = 16
            row += 1
        return row

    # -- growth / variance --------------------------------------------------

    def _growth(self, ws, growth_table, row):
        if not growth_table:
            return row
        is_variance = any("variance" in str(k).lower() for k in growth_table[0].keys())
        row = self._section_header(ws, "VARIANCE ANALYSIS" if is_variance else "GROWTH ANALYSIS", row)
        columns = [c for c in growth_table[0].keys() if c != "direction"]
        row = self._table_header(ws, columns, row)

        border = S.box(S.BORDER_COLOR, "thin")
        for record in growth_table:
            base_bg = S.ROW_ALT if row % 2 == 0 else S.ROW_MAIN
            for c_index, col in enumerate(columns, start=1):
                value = record.get(col)
                lowered = str(col).lower()
                signed = ("growth" in lowered or "variance" in lowered or "→" in str(col)) and \
                    isinstance(value, (int, float)) and not isinstance(value, bool)
                if signed and value > 0:
                    bg, color = S.POSITIVE_BG, S.POSITIVE_TEXT
                elif signed and value < 0:
                    bg, color = S.NEGATIVE_BG, S.NEGATIVE_TEXT
                else:
                    bg, color = base_bg, S.TEXT_DARK
                S.style_cell(ws.cell(row=row, column=c_index), value=self._fmt(value),
                             font_=S.font(10, color=color), fill_=S.fill(bg),
                             align=S.RIGHT if isinstance(value, (int, float)) and not isinstance(value, bool) else S.LEFT,
                             border=border, number_format=S.number_format_for(col, value))
            ws.row_dimensions[row].height = 16
            row += 1
        return row

    # -- insights -----------------------------------------------------------

    def _insights(self, ws, insights, row):
        if not insights:
            return row
        row = self._section_header(ws, "KEY INSIGHTS", row)
        for text in insights:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=SECTION_SPAN)
            S.style_cell(ws.cell(row=row, column=1), value=f"•  {text}",
                         font_=S.font(10, color=S.TEXT_DARK), fill_=S.fill(S.ROW_MAIN), align=S.LEFT)
            ws.row_dimensions[row].height = 18
            row += 1
        return row

    # -- shared -------------------------------------------------------------

    def _table_header(self, ws, columns, row):
        border = S.header_border()
        for c_index, col in enumerate(columns, start=1):
            S.style_cell(ws.cell(row=row, column=c_index), value=humanize_header(col),
                         font_=S.font(10, bold=True, color=S.HEADER_FONT), fill_=S.fill(S.HEADER_BG),
                         align=S.CENTER, border=border)
        ws.row_dimensions[row].height = 18
        return row + 1

    @staticmethod
    def _fmt(value):
        if value is None:
            return "-"
        if isinstance(value, bool):
            return value
        if isinstance(value, float):
            return round(value, 2)
        return value
