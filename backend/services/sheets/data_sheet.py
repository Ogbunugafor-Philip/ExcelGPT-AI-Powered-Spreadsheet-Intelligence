"""Data sheet — the PRIMARY computed result, with number formats, filters and rules.

The question-driven column/row selection happens upstream in the OutputPackager
(growth → variance → rankings → grouped → raw). This builder renders whatever
primary table it is handed, applying the correct number format to every cell and
the world-class light theme (FIX 2 & 6).
"""

from __future__ import annotations

import re
from typing import Any

import pandas as pd
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

from . import styles as S

_RULE_OPS = {
    "<": "lessThan", "<=": "lessThanOrEqual", ">": "greaterThan",
    ">=": "greaterThanOrEqual", "==": "equal", "=": "equal", "!=": "notEqual",
}

_ARROWS = {"up": ("▲", S.POSITIVE_TEXT), "down": ("▼", S.NEGATIVE_TEXT), "neutral": ("▬", S.TEXT_MUTED)}
_HEADER_H = 22
_ROW_H = 16


class DataSheet:
    def build(self, ws, data_sheet: dict[str, Any], output: dict[str, Any] | None = None) -> None:
        S.hide_gridlines(ws)
        ws.sheet_properties.tabColor = S.SECTION_HEADER

        columns = [S.column_name(c) for c in data_sheet.get("columns", [])]
        rows = self._normalise_rows(data_sheet.get("rows", []), columns)
        if not columns:
            S.style_cell(ws["A1"], value="No data available.", font_=S.font(11, italic=True, color=S.TEXT_MUTED))
            return

        col_types = {
            col: S.detect_type(col, [row[idx] for row in rows if idx < len(row)])
            for idx, col in enumerate(columns)
        }

        # Header row
        for idx, col in enumerate(columns, start=1):
            S.style_cell(ws.cell(row=1, column=idx), value=col, font_=S.font(11, bold=True, color=S.HEADER_FONT),
                         fill_=S.fill(S.HEADER_BG), align=S.CENTER, border=S.header_border())
        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = _HEADER_H

        last_col_letter = get_column_letter(len(columns))
        last_row = len(rows) + 1
        ws.auto_filter.ref = f"A1:{last_col_letter}{max(last_row, 1)}"

        # Data rows
        border = S.box(S.BORDER_COLOR, "thin")
        for r_index, row in enumerate(rows, start=2):
            bg = S.ROW_ALT if r_index % 2 == 0 else S.ROW_MAIN
            ws.row_dimensions[r_index].height = _ROW_H
            for c_index, col in enumerate(columns, start=1):
                value = row[c_index - 1] if c_index - 1 < len(row) else None
                cell = ws.cell(row=r_index, column=c_index)
                self._write_value(cell, value, col, col_types[col], bg, border)

        S.autofit_columns(ws, columns, rows)
        self._apply_conditional_formatting(ws, columns, data_sheet.get("conditional_formatting", []), last_row)
        self._tint_signed_columns(ws, columns, col_types, last_row)
        S.paint_canvas(ws, max_row=last_row, max_col=len(columns), color=S.ROW_MAIN)

    # -- helpers ------------------------------------------------------------

    def _normalise_rows(self, rows: list[Any], columns: list[str]) -> list[list[Any]]:
        normalised = []
        for row in rows:
            if isinstance(row, dict):
                normalised.append([row.get(col) for col in columns])
            elif isinstance(row, (list, tuple)):
                normalised.append(list(row))
            else:
                normalised.append([row])
        return normalised

    def _write_value(self, cell, value, col_name, col_type, bg, border):
        if value is None or (isinstance(value, str) and value.strip() == ""):
            S.style_cell(cell, value="-", font_=S.font(10, color=S.TEXT_MUTED), fill_=S.fill(bg),
                         align=S.CENTER, border=border)
            return

        # Direction columns render as a coloured arrow glyph.
        if str(col_name).lower() in ("direction", "trend") and str(value).lower() in _ARROWS:
            arrow, color = _ARROWS[str(value).lower()]
            S.style_cell(cell, value=arrow, font_=S.font(11, bold=True, color=color), fill_=S.fill(bg),
                         align=S.CENTER, border=border)
            return

        # Status columns: green fill for "Above Target", red for "Below Target".
        if "status" in str(col_name).lower() and isinstance(value, str):
            low = value.lower()
            if "above" in low or "over" in low:
                fill_c, text_c = S.POSITIVE_BG, S.POSITIVE_TEXT
            elif "below" in low or "under" in low:
                fill_c, text_c = S.NEGATIVE_BG, S.NEGATIVE_TEXT
            else:
                fill_c, text_c = bg, S.TEXT_DARK
            S.style_cell(cell, value=value, font_=S.font(10, bold=True, color=text_c), fill_=S.fill(fill_c),
                         align=S.CENTER, border=border)
            return

        number_format = None
        if col_type == "currency":
            cell.value = self._num(value)
            number_format = S.number_format_for(col_name, cell.value) or S.NAIRA_FORMAT
            align = S.RIGHT
        elif col_type == "percentage":
            cell.value = self._num(value)
            number_format = S.PERCENT_FORMAT
            align = S.RIGHT
        elif col_type == "date":
            parsed = pd.to_datetime(value, errors="coerce")
            if pd.notna(parsed):
                cell.value = parsed.to_pydatetime()
                number_format = S.DATE_FORMAT
            else:
                cell.value = value
            align = S.LEFT
        elif col_type == "number":
            cell.value = self._num(value)
            number_format = S.number_format_for(col_name, cell.value)
            align = S.RIGHT
        else:
            cell.value = value
            align = S.LEFT

        S.style_cell(cell, font_=S.font(10, color=S.TEXT_DARK), fill_=S.fill(bg), align=align,
                     border=border, number_format=number_format)

    @staticmethod
    def _num(value):
        if isinstance(value, bool):
            return value
        if isinstance(value, (int, float)):
            return value
        try:
            return float(str(value).replace(",", "").replace("₦", "").replace("%", "").strip())
        except (TypeError, ValueError):
            return value

    def _tint_signed_columns(self, ws, columns, col_types, last_row):
        """Green for positive, red for negative on growth/variance numeric columns."""
        if last_row < 2:
            return
        for idx, col in enumerate(columns):
            lowered = str(col).lower()
            if not ("growth" in lowered or "variance" in lowered or "change" in lowered):
                continue
            if col_types.get(col) not in ("percentage", "number", "currency"):
                continue
            letter = get_column_letter(idx + 1)
            rng = f"{letter}2:{letter}{last_row}"
            ws.conditional_formatting.add(rng, CellIsRule(operator="greaterThan", formula=["0"],
                                                          fill=S.fill(S.POSITIVE_BG), font=S.font(10, color=S.POSITIVE_TEXT)))
            ws.conditional_formatting.add(rng, CellIsRule(operator="lessThan", formula=["0"],
                                                          fill=S.fill(S.NEGATIVE_BG), font=S.font(10, color=S.NEGATIVE_TEXT)))

    def _apply_conditional_formatting(self, ws, columns, rules, last_row):
        if last_row < 2:
            return
        index_by_name = {name: i for i, name in enumerate(columns)}
        for rule in rules or []:
            column = rule.get("column")
            if column not in index_by_name:
                continue
            letter = get_column_letter(index_by_name[column] + 1)
            cell_range = f"{letter}2:{letter}{last_row}"
            operator, formula = self._parse_rule(rule.get("rule", ""))
            if operator is None:
                continue
            color = (rule.get("color") or S.NEGATIVE_BG).lstrip("#")
            ws.conditional_formatting.add(
                cell_range,
                CellIsRule(operator=operator, formula=[formula], fill=S.fill(color),
                           font=S.font(10, bold=True, color=S.NEGATIVE_TEXT)),
            )

    @staticmethod
    def _parse_rule(expression: str):
        match = re.search(r"(<=|>=|==|!=|=|<|>)\s*(-?\d+(?:\.\d+)?)", str(expression))
        if not match:
            return None, None
        return _RULE_OPS.get(match.group(1)), match.group(2)
