"""Charts sheet — embeds the matplotlib-rendered PNGs, one per row, stacked."""

from __future__ import annotations

import logging
import os
from typing import Any

from openpyxl.drawing.image import Image as XLImage

from . import styles as S

logger = logging.getLogger(__name__)

CHART_W = 500
CHART_H = 300
ROW_SPAN = 17          # rows each chart block occupies (title + image)


class ChartsSheet:
    def build(self, ws, charts: list[dict[str, Any]]) -> None:
        S.hide_gridlines(ws)
        ws.sheet_properties.tabColor = S.AMBER
        S.set_column_widths(ws, 18, first=1, last=10)

        if not charts:
            S.style_cell(ws["A1"], value="No charts generated for this report.",
                         font_=S.font(11, italic=True, color=S.TEXT_MUTED))
            S.paint_canvas(ws, max_row=2, max_col=10, color=S.ROW_MAIN)
            return

        max_row = 1
        for index, chart in enumerate(charts):
            title_row = 1 + index * ROW_SPAN
            image_row = title_row + 1

            # Bold chart title above the image.
            ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=6)
            S.style_cell(ws.cell(row=title_row, column=1), value=chart.get("title", "Chart"),
                         font_=S.font(13, bold=True, color=S.SECTION_HEADER), align=S.LEFT)
            ws.row_dimensions[title_row].height = 20

            image_path = chart.get("image_path")
            anchor = f"A{image_row}"
            if image_path and os.path.exists(image_path):
                try:
                    img = XLImage(image_path)
                    img.width = CHART_W
                    img.height = CHART_H
                    ws.add_image(img, anchor)
                except Exception as exc:  # noqa: BLE001
                    logger.warning("Failed to embed chart image %s: %s", image_path, exc)
                    self._placeholder(ws, image_row, f"Chart could not be embedded ({exc})")
            else:
                logger.warning("Chart image not found, skipping embed: %s", image_path)
                self._placeholder(ws, image_row, f"Chart image not found: {image_path}")

            ws.row_dimensions[image_row].height = CHART_H * 0.75  # px -> points approximation
            max_row = max(max_row, title_row + ROW_SPAN)

        S.paint_canvas(ws, max_row=max_row, max_col=10, color=S.ROW_MAIN)

    def _placeholder(self, ws, row, text):
        S.style_cell(ws.cell(row=row, column=1), value=text,
                     font_=S.font(11, italic=True, color=S.NEGATIVE_TEXT), fill_=S.fill(S.NEGATIVE_BG), align=S.LEFT)
