"""
ExcelBuilder — assemble the multi-sheet .xlsx workbook from a ComputationOutput.

Each output sheet is delegated to a dedicated builder under services/sheets/.
Empty sheets are skipped. The finished workbook is written under the session's
output directory and the path is returned for the /download endpoint to stream.
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, NamedStyle, PatternFill
from openpyxl.utils import get_column_letter

import config
from schemas.computation_schema import ComputationOutput

from .sheets import styles as S
from .sheets.analysis_sheet import AnalysisSheet
from .sheets.charts_sheet import ChartsSheet
from .sheets.data_sheet import DataSheet
from .sheets.executive_summary_sheet import ExecutiveSummarySheet
from .sheets.forecast_sheet import ForecastSheet


class ExcelBuilder:
    def __init__(self) -> None:
        self.executive = ExecutiveSummarySheet()
        self.data = DataSheet()
        self.analysis = AnalysisSheet()
        self.charts = ChartsSheet()
        self.forecast = ForecastSheet()

    def build(self, output: ComputationOutput, session_id: str) -> str:
        data = output.model_dump() if isinstance(output, ComputationOutput) else dict(output)

        wb = Workbook()
        wb.remove(wb.active)  # drop the default empty sheet
        self.apply_global_styles(wb)

        # Executive Summary — always present.
        ws = wb.create_sheet("Executive Summary")
        self.executive.build(ws, data.get("executive_summary", {}))

        # Data — the PRIMARY question-driven result; skip when there are no columns.
        data_sheet = data.get("data_sheet", {})
        if data_sheet.get("columns"):
            self.data.build(wb.create_sheet("Data"), data_sheet, output=data)

        # Analysis — skip when nothing derived.
        analysis = data.get("analysis_sheet", {})
        if analysis.get("metrics") or analysis.get("rankings") or analysis.get("growth_table") or analysis.get("insights"):
            self.analysis.build(wb.create_sheet("Analysis"), analysis)

        # Charts — skip when none rendered.
        charts = data.get("charts", [])
        if charts:
            self.charts.build(wb.create_sheet("Charts"), charts)

        # Forecast — skip when no series.
        forecast = data.get("forecast_sheet", {})
        if ForecastSheet.has_content(forecast):
            self.forecast.build(wb.create_sheet("Forecast"), forecast)

        output_dir = Path(config.UPLOAD_DIR) / session_id / "output"
        output_dir.mkdir(parents=True, exist_ok=True)
        file_path = output_dir / "report.xlsx"
        wb.save(file_path)
        return str(file_path)

    # -- combined "Download All" workbook -----------------------------------

    def build_combined(self, items: list[tuple[str, dict[str, Any]]], session_id: str) -> str:
        """Package several insights into ONE workbook — one sheet per insight.

        ``items`` is a list of (label, computation_output_dict). The label
        becomes the sheet name (first 31 chars, sanitised + de-duplicated).
        """
        wb = Workbook()
        wb.remove(wb.active)
        self.apply_global_styles(wb)

        used: set[str] = set()
        for index, (label, data) in enumerate(items, start=1):
            sheet_name = self._safe_sheet_name(label or f"Insight {index}", used)
            ws = wb.create_sheet(sheet_name)
            self._render_insight_sheet(ws, data, label or sheet_name)

        if not wb.sheetnames:  # defensive — never save an empty workbook
            self._render_insight_sheet(wb.create_sheet("Insights"), {}, "Insights")

        output_dir = Path(config.UPLOAD_DIR) / session_id / "output"
        output_dir.mkdir(parents=True, exist_ok=True)
        file_path = output_dir / "insights_combined.xlsx"
        wb.save(file_path)
        return str(file_path)

    def _render_insight_sheet(self, ws, data: dict[str, Any], label: str) -> None:
        """Render one insight. Prefer the question-driven Data table; otherwise
        fall back to a simple KPI / metrics summary so the sheet is never blank."""
        data_sheet = (data or {}).get("data_sheet", {}) or {}
        if data_sheet.get("columns"):
            self.data.build(ws, data_sheet, output=data)
            return
        self._render_summary_sheet(ws, data or {}, label)

    def _render_summary_sheet(self, ws, data: dict[str, Any], label: str) -> None:
        S.hide_gridlines(ws)
        summary = data.get("executive_summary", {}) or {}
        analysis = data.get("analysis_sheet", {}) or {}
        title = summary.get("title") or label

        ws.merge_cells("A1:D1")
        S.style_cell(ws["A1"], value=title, font_=S.font(15, bold=True, color=S.HEADER_FONT),
                     fill_=S.fill(S.HEADER_BG), align=S.LEFT)
        ws.row_dimensions[1].height = 26
        row = 3

        cards = summary.get("kpi_cards", []) or []
        if cards:
            S.style_cell(ws.cell(row=row, column=1), value="KEY FIGURES",
                         font_=S.font(11, bold=True, color=S.SECTION_FONT), fill_=S.fill(S.SECTION_HEADER), align=S.LEFT)
            for col in range(2, 4):
                ws.cell(row=row, column=col).fill = S.fill(S.SECTION_HEADER)
            row += 1
            for i, card in enumerate(cards):
                bg = S.ROW_ALT if i % 2 == 0 else S.ROW_MAIN
                S.style_cell(ws.cell(row=row, column=1), value=card.get("label", ""),
                             font_=S.font(10, bold=True, color=S.TEXT_DARK), fill_=S.fill(bg), align=S.LEFT)
                S.style_cell(ws.cell(row=row, column=2), value=card.get("value", ""),
                             font_=S.font(10, color=S.TEXT_DARK), fill_=S.fill(bg), align=S.RIGHT)
                S.style_cell(ws.cell(row=row, column=3), value=card.get("change", ""),
                             font_=S.font(10, color=S.TEXT_MUTED), fill_=S.fill(bg), align=S.LEFT)
                row += 1
            row += 1

        for text in analysis.get("insights", []) or []:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            S.style_cell(ws.cell(row=row, column=1), value=f"•  {text}",
                         font_=S.font(10, color=S.TEXT_DARK), fill_=S.fill(S.ROW_MAIN), align=S.LEFT)
            row += 1

        if not cards and not (analysis.get("insights")):
            S.style_cell(ws.cell(row=row, column=1), value="See the in-app insight for the full breakdown.",
                         font_=S.font(10, italic=True, color=S.TEXT_MUTED))

        for col in range(1, 4):
            ws.column_dimensions[get_column_letter(col)].width = 32 if col == 1 else 22

    @staticmethod
    def _safe_sheet_name(label: str, used: set[str]) -> str:
        """Excel sheet names: ≤31 chars, none of : \\ / ? * [ ], and unique."""
        name = re.sub(r"[:\\/?*\[\]]", " ", str(label)).strip() or "Insight"
        name = name[:31]
        candidate = name
        suffix = 2
        while candidate.lower() in used or not candidate:
            tail = f" {suffix}"
            candidate = (name[: 31 - len(tail)] + tail).strip()
            suffix += 1
        used.add(candidate.lower())
        return candidate

    def apply_global_styles(self, wb: Workbook) -> None:
        """Register the ExcelGPT named styles and set a sane default font."""
        # Default font for any cell that doesn't override it.
        wb._default_font = Font(name=S.DEFAULT_FONT, size=11)  # noqa: SLF001 — openpyxl has no public setter

        styles: dict[str, dict[str, Any]] = {
            "header_style": dict(font=Font(name="Calibri", size=11, bold=True, color=S.WHITE),
                                 fill=PatternFill("solid", fgColor=S.BLUE_ELECTRIC),
                                 alignment=Alignment(horizontal="center", vertical="center")),
            "data_style": dict(font=Font(name="Calibri", size=10, color=S.WHITE),
                               fill=PatternFill("solid", fgColor=S.NAVY)),
            "currency_style": dict(font=Font(name="Calibri", size=10, color=S.WHITE),
                                   fill=PatternFill("solid", fgColor=S.NAVY),
                                   alignment=Alignment(horizontal="right", vertical="center"),
                                   number_format=S.CURRENCY_FORMAT),
            "kpi_value_style": dict(font=Font(name="Calibri", size=18, bold=True, color=S.WHITE),
                                    fill=PatternFill("solid", fgColor=S.NAVY_LIGHT),
                                    alignment=Alignment(horizontal="center", vertical="center")),
            "section_header_style": dict(font=Font(name="Calibri", size=13, bold=True, color=S.BLUE_ELECTRIC)),
        }

        for name, spec in styles.items():
            if name in wb.named_styles:
                continue
            style = NamedStyle(name=name)
            style.font = spec["font"]
            if "fill" in spec:
                style.fill = spec["fill"]
            if "alignment" in spec:
                style.alignment = spec["alignment"]
            if "number_format" in spec:
                style.number_format = spec["number_format"]
            wb.add_named_style(style)
